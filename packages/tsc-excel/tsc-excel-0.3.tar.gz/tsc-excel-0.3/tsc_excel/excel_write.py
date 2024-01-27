# 存在简单依赖第三方包的类和方法
import os
import re
from pprint import pprint
from copy import deepcopy
from typing import Dict, Union, Hashable, List, Optional, Callable, Literal
from tsc_base import dict_to_pair, pair_to_dict, merge_dict, get, any_value_dict, recur_opt_any
import xlwt
import openpyxl
from tqdm import tqdm
from openpyxl.worksheet.worksheet import Worksheet as openpyxl_Worksheet
from openpyxl.workbook.workbook import Workbook as openpyxl_Workbook
from xlwt.Workbook import Workbook as xlwt_Workbook
from xlwt.Worksheet import Worksheet as xlwt_Worksheet

NestedDict = Dict[Hashable, Union[int, 'NestedDict']]


class Excel:
    # 用于 openpyxl.utils.exceptions.IllegalCharacterError
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    @staticmethod
    def get_excel_th(
        th_D: dict, 
        horizontal: bool = True, 
        start_r: int = 0, 
        start_c: int = 0, 
        rectangle: str = '$rectangle', 
        sort_key: Optional[dict] = None, 
        wild_card: str = '$*v', 
        max_deep: Optional[int] = None):
        """递归将dict转换为具有单元格位置坐标的excel表头, 用于 xlwt 等, 行列值从0开始

        Args:
            th_D (dict): 会把所有key拿出来作为表头, 然后value中增加 {rectangle:(开始行,结束行,开始列,结束列)}, 注意会覆盖原始value
            horizontal (bool, optional): 是否横向展开表头
            start_r (int, optional): 表头占据的开始行, 通常xlwt从0开始, openpyxl从1开始, 不过excel_add_sheet中进行了统一，所以这里都是从0开始
                行列是相对的，horizontal=False 这个相对于整个表而言就是列了。但是存到4元组后就是绝对值了
            start_c (int, optional): 表头占据的开始列
            rectangle (str, optional): 作为存储每个单元格的4角坐标的key, 也用于sort_key中的func, 不能与th_D中任意key一样
            sort_key (dict, optional): key:{rectangle:{'f':f(k,v),'r':bool},..}; 用于排序同一个字典中的k, 默认不排序
                f: 函数f输入key下面value中的k和v, 返回排序值, k会按照排序值排序, 例如 lambda t:t[0]
                r (bool, optional): 在 f 存在的情况下, 是否倒序排序, 不存在默认False
            wild_card (str, optional): 用于 sort_key 中匹配key的通配符, 如果在sort_key中找不到key则进入wild_card中的value, 不能与th_D中任意key一样
            max_deep (int, optional): 用于递归确定深度, 不能修改

        Returns:
            c, max_deep: int,int; 总占据列宽, 总占据行数
        """
        if max_deep is None:
            x = list(dict_to_pair(th_D))
            for i in x:
                assert len({rectangle, wild_card} & set(i[0])) == 0, 'th_D 中存在 rectangle/wild_card 标记, 请进行修改:' + str(i)
            max_deep = max([len(i[0]) for i in x])
        sort_key = {} if sort_key is None else sort_key
        if rectangle in sort_key:  # 排序
            par = sorted(th_D.items(), key=sort_key[rectangle]['f'], reverse=bool(sort_key[rectangle].get('r')))
        else:
            par = th_D.items()
        c = 0  # 深度优先得到上层的列宽
        for k, v in par:
            start_c_last = start_c + c  # 随着循环, 列的起始点在变化, 行的起始点不变
            r = max_deep
            if isinstance(v, dict):  # 随着递归, 行每次向下移动一行
                cc = Excel.get_excel_th(
                    v, horizontal, start_r+1, start_c_last, rectangle,
                    sort_key.get(k) or sort_key.get(wild_card), wild_card, max_deep-1)[0]
                if cc:
                    c += cc
                    r = 1  # 下面还要分之则只能占一行
                else:
                    c += 1  # 空字典
            else:
                c += 1  # 只占一列
                th_D[k] = v = {}
            v[rectangle] = (start_r, start_r+r-1, start_c_last, start_c+c-1)
            v[rectangle] = v[rectangle] if horizontal else (*v[rectangle][2:], *v[rectangle][:2])
        return c, max_deep

    @staticmethod
    def get_excel_table(
        doc_L: List[dict], 
        ignore_th_f: Callable = lambda t: t, 
        td_f: Callable = lambda t: t[1] if isinstance(t[1], (int, float)) else str(t[1]),
        ordered: bool = True,
        horizontal: bool = True, 
        rectangle: str = '$rectangle',
        **kw):
        """将多个dict转换为excel表格的单元格坐标, 用于生成excel表格

        Args:
            doc_L (list): [{..},..]; 被转换的dict列表
            ignore_th_f (func, optional): 输入 (key_L,value) 返回 (key_L,value) or None, None表示丢弃key_L这个表头
                用于修剪层次过深的表头, 这时候可能需要 td_f 优化dict在单元格中的展示形式
            td_f (func, optional): 用于优化单元格中的值, value是func, 输入 (key_L,value) 返回优化的展示 value
                小心处理每一种值格式, 因为 value 不能是 dict,list,tuple,set 等类型, 否则可能导致 excel 写入出错
            ordered (bool, optional): 是否保留 doc_L 中字段间的顺序(不是doc的顺序)，使用后速度会慢一些
            horizontal (bool, optional): 见 get_excel_th
            rectangle (str, optional): 见 get_excel_th
            **kw: 其他参数见 get_excel_th

        Returns:
            dict, list, list: th_D 的一个例子:
                {'inf_jcr': {'$rectangle': (0, 0, 1, 33),
                'Open Access': {'$rectangle': (1, 5, 29, 29)},
                '期刊分区': {'$rectangle': (1, 1, 1, 27),
                        'JCR分区': {'$rectangle': (2, 5, 27, 27)},}}
        """
        # 表头
        pair = []
        for p in dict_to_pair(merge_dict(doc_L, ordered=ordered)):
            p = ignore_th_f(p)
            if p is not None:
                pair.append(p)
        th_D = pair_to_dict(pair)
        Excel.get_excel_th(th_D, horizontal=horizontal, rectangle=rectangle, **kw)
        th_L = [(i[1], i[0][-2]) for i in dict_to_pair(th_D)]  # 与 td_L 格式一致
        # 表中值
        td_L = []  # [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 与doc_L顺序一致
        for no, doc in enumerate(doc_L):
            coor_v_D = {}  # {(开始行,结束行,开始列,结束列):单元格值,..}
            for p in dict_to_pair(doc):
                p = ignore_th_f(p)  # (key_L,value); 用于获取行列位置, 位置靠前了值也会变
                # 检查这个值是否满足要求
                if p is None:
                    continue
                th = get(p[0], th_D)  # 对应的表头坐标
                if len(th) > 1:  # 没有到叶子结点, 说明doc存在大段的空
                    continue
                # 获取坐标
                _, r, _, c = th[rectangle]  # 右下角单元格坐标
                r, c = (r + no + 1, c) if horizontal else (r, c + no + 1)
                coor = (r, r, c, c)  # (开始行,结束行,开始列,结束列)
                # 保存坐标与单元格值
                if coor not in coor_v_D:  # 防止重复
                    coor_v_D[coor] = td_f((p[0], get(p[0], doc)))
            td_L.append(list(coor_v_D.items()))
        return th_D, th_L, td_L

    @staticmethod
    def excel_add_sheet(
        workbook: Union[openpyxl_Workbook, xlwt_Workbook],
        name: str, 
        th_L: list, 
        td_L: list, 
        package,
        index: Optional[int] = 0, 
        save_path: Optional[str] = None, 
        tqdm_f = None,
        auto_adjust_width: bool = False,
    ):
        """使用行列坐标 写入一页 excel 表格

        Args:
            workbook (obj): openpyxl.Workbook() or xlwt.Workbook(encoding='utf8')
            name (str): sheet 名称
            th_L (list): [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 行列编号从0开始, 0行就是第一行
            td_L (list): [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 行列编号从0开始, 0行就是第一行
            package : (xlwt>=1.3.0 import xlwt) or (openpyxl>=3.0.9 import openpyxl)
            index (int, optional): 只用于 openpyxl, 表示插入的 sheet 的位置, xlwt 只是追加
            save_path (str, optional): 写入文件的路径, 会自动添加后缀名
            tqdm_f : tqdm>=4.62.3 from tqdm import tqdm
            auto_adjust_width (bool, optional): 是否自动调整非行合并的单元格列宽以适应文本长度, 只有 openpyxl 支持

        Returns:
            workbook
        """
        if 'openpyxl' in str(type(workbook)):  # openpyxl - xlsx (打开效率更高)
            worksheet: openpyxl_Worksheet = workbook.create_sheet(name, index)
            # 写表头
            for coor, v in th_L:
                coor = [i+1 for i in coor]  # openpyxl 从1开始
                worksheet.merge_cells(start_row=coor[0], start_column=coor[2], end_row=coor[1], end_column=coor[3])
                cell = worksheet.cell(coor[0], coor[2])
                cell.value = Excel.ILLEGAL_CHARACTERS_RE.sub('', v) if type(v) == str else v
                cell.alignment = package.styles.Alignment(horizontal='center', vertical='center')
                cell.font = package.styles.Font(bold=True)
            # 单元格值
            par = tqdm_f(td_L, f'{name}-写入表格(openpyxl)') if tqdm_f else td_L
            for i in par:
                for coor, v in i:
                    cell = worksheet.cell(coor[0]+1, coor[2]+1)  # openpyxl 从1开始
                    cell.value = Excel.ILLEGAL_CHARACTERS_RE.sub('', v) if type(v) == str else v
                    cell.alignment = package.styles.Alignment(horizontal='center', vertical='center')
            if save_path:
                if not save_path.endswith('.xlsx'):
                    save_path = save_path + '.xlsx'
            else:
                save_path = None
            if auto_adjust_width:
                column_widths = {}
                for row in worksheet.iter_rows():  # 遍历所有行
                    for i, cell in enumerate(row):  # 遍历行中的所有单元格
                        if (  # 行的合并单元格跳过
                            isinstance(cell, package.cell.cell.MergedCell) or
                            i < len(row) - 1 and isinstance(row[i + 1], package.cell.cell.MergedCell)
                        ):
                            continue
                        if cell.value:
                            column_letter = cell.column_letter
                            current_width = column_widths.get(column_letter, 0)
                            new_width = len(str(cell.value))
                            column_widths[column_letter] = max(current_width, new_width)
                for column_letter, width in column_widths.items():  # 设置列宽
                    worksheet.column_dimensions[column_letter].width = width + 2
        else:  # xlwt - xls (写入快6,7倍)
            worksheet: xlwt_Worksheet = workbook.add_sheet(name)
            # 单元格值
            style = package.XFStyle()
            alignment = package.Alignment()
            alignment.horz = package.Alignment.HORZ_CENTER
            alignment.vert = package.Alignment.VERT_CENTER
            style.alignment = alignment
            par = tqdm_f(td_L, f'{name}-写入表格(xlwt)') if tqdm_f else td_L
            for i in par:
                for coor, v in i:
                    worksheet.write_merge(*coor, v, style)
            # 写表头
            font = package.Font()
            font.bold = True
            style.font = font
            for coor, v in th_L:
                worksheet.write_merge(*coor, v, style)
            if save_path:
                if not save_path.endswith('.xls'):
                    save_path = save_path + '.xls'
            else:
                save_path = None
        if save_path:
            if os.path.dirname(save_path) and not os.path.exists(os.path.dirname(save_path)):
                os.makedirs(os.path.dirname(save_path))
            workbook.save(save_path)
        return worksheet

    @classmethod
    def get_excel_table_with_side(cls, doc_L: List[dict], side_th: NestedDict = None, **kwargs) -> dict:
        """将多个dict转换为excel表格的单元格坐标, 用于生成excel表格, 侧边栏表头
        最终用于 excel_add_sheet 的方式：
        th_L = ret['th_L'] + ret['side_th_L']
        td_L = ret['td_L']
        注意：主表头的最后一行一定不是合并单元格，侧边栏表头的最后一行是不是都行

        Args:
            doc_L (List[dict]): [{..},..]; 被转换的dict列表
            side_th (NestedDict, optional): 侧边栏表头, 会在表格左边或上边生成一个表格, 用于存储一些额外信息
            **kwargs: 其他参数见 get_excel_table

        Returns:
            dict: 参考 get_excel_table
        """
        if side_th:
            doc_num = any_value_dict(side_th, lambda v1, v2: v1 + v2, 0, True)  # 有非int会报错
            assert doc_num > 0, 'side_th 中没有文档!'
            assert doc_num == len(doc_L), f'从 side_th 得到的文档数 {doc_num} 与 doc_L 中的 {len(doc_L)} 不一致'
            side_th_high = max(len(i[0]) for i in dict_to_pair(side_th))  # 侧边栏表头的最大深度
            side_th_ext = deepcopy(side_th)
            recur_opt_any(side_th_ext, lambda k, t, v: ({i: i for i in range(v)}, 0, 1) if isinstance(v, int) else (v, 0, 0))
            # 表头参数
            kwargs['ordered'] = True  # doc_L 顺序不能变
            side_kw = kwargs.copy()
            side_kw['horizontal'] = not kwargs.get('horizontal', True)  # 侧边栏表头的横竖是相反的
            kwargs['start_c'] = kwargs.get('start_c', 0) + side_th_high  # 因为 start_c 是相对的，永远是列
            # 核心表格计算
            th_D, th_L, td_L = cls.get_excel_table(doc_L, **kwargs)
            th_high = max(len(i[0]) for i in dict_to_pair(th_D)) - 1  # 核心表格的最大深度
            # 侧边栏表格计算
            side_kw['start_c'] = side_kw.get('start_c', 0) + th_high
            side_th_D, side_th_L, side_td_L = cls.get_excel_table([side_th_ext], **side_kw)
            
            # 去除最后一行，由side_th_ext引入的辅助行
            # 删除叶子
            recur_opt_any(side_th_D, lambda k, t, v: (v, 1, 0) if isinstance(v, dict) and len(v) <= 1 else (v, 0, 0))
            # 确定补全行或列的位置，防止合并单元格不全
            if side_kw['horizontal']:
                complete_no = 1  # 补全行
            else:
                complete_no = 3  # 补全列
            max_no = side_kw.get('start_r', 0) + side_th_high - 1
            # 补全行或列
            rectangle = kwargs.get('rectangle', '$rectangle')
            
            def opt(keys, types, value):
                if isinstance(value, dict) and len(value) == 1:
                    coor = value[rectangle]
                    value[rectangle] = coor[:complete_no] + (max_no,) + coor[complete_no+1:]
                    return value, 0, 1
                return value, 0, 0
            recur_opt_any(side_th_D, opt)
            # 重新生成侧边栏表头坐标
            side_th_L = [(i[1], i[0][-2]) for i in dict_to_pair(side_th_D)]
        else:
            th_D, th_L, td_L = cls.get_excel_table(doc_L, **kwargs)
            side_th_D, side_th_L, side_td_L = {}, [], []
        return {
            'th_D': th_D,
            'th_L': th_L,
            'td_L': td_L,
            'side_th_D': side_th_D,
            'side_th_L': side_th_L,
        }


if __name__ == '__main__':
    # Excel
    print('=' * 10, 'Excel')
    doc_L = [{
        'a111111你': {'b11': i, 'b2': 2, 'b3': {'c1': 1, 'c22222': False}},
        'a2': '123123123',
        'a3': {},
        'a4': None,
        'a5': [],
    } for i in range(10)]
    openpyxl_wb = openpyxl.Workbook()
    xlwt_wb = xlwt.Workbook(encoding='utf8')
    ignore_th_f = lambda t: None if t[1] == {} or t[1] == [] else t
    
    for horizontal, name in [(True, '横向'), (False, '纵向')]:
        th_D, th_L, td_L = Excel.get_excel_table(doc_L, ignore_th_f=ignore_th_f, horizontal=horizontal)
        print('th_D:')
        pprint(th_D)
        Excel.excel_add_sheet(openpyxl_wb, f'{name}', th_L, td_L, openpyxl, 0, 'test/openpyxl', tqdm, auto_adjust_width=True)
        Excel.excel_add_sheet(xlwt_wb, f'{name}', th_L, td_L, xlwt, 0, 'test/xlwt', tqdm, auto_adjust_width=True)
        
        print('=' * 5, 'add side:')
        ret = Excel.get_excel_table_with_side(doc_L, side_th={
            'test': {'a': 1, 'b'*10: 2, 'c': {'d': 1, 'e': len(doc_L) - 4}},
        }, ignore_th_f=ignore_th_f, horizontal=horizontal)
        print('side_th_D:')
        pprint(ret['side_th_D'])
        th_L = ret['th_L'] + ret['side_th_L']
        td_L = ret['td_L']
        Excel.excel_add_sheet(openpyxl_wb, f'{name}2', th_L, td_L, openpyxl, 0, 'test/openpyxl', tqdm, auto_adjust_width=True)
        Excel.excel_add_sheet(xlwt_wb, f'{name}2', th_L, td_L, xlwt, 0, 'test/xlwt', tqdm, auto_adjust_width=True)
        
        print()
