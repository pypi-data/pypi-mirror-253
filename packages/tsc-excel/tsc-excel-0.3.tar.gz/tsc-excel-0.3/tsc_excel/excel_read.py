from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.merge import MergedCellRange
from typing import List, Generator, Tuple, Union, Optional, Dict, Hashable, TypedDict
from tsc_base import pair_to_dict, dict_to_pair
from pprint import pprint

NestedDict = Dict[Hashable, Union[int, 'NestedDict']]


class XlsxReader:
    @staticmethod
    def find_merged_cell(sheet: Worksheet, cell: Cell) -> Union[MergedCellRange, None]:
        """查找单元格是否是合并单元格的一部分

        Args:
            sheet (Worksheet): 工作表
            cell (Cell): 单元格

        Returns:
            Union[MergedCellRange, None]: 如果是合并单元格的一部分，返回合并单元格的范围，否则返回 None
        """
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return merged_range
        return None

    @classmethod
    def get_table_header(
        cls,
        min_col: Union[int, str],
        min_row: int, 
        max_col: Union[int, str],
        max_row: int, 
        sheet: Worksheet, 
        horizontal: bool = True, 
        root_list: List[list] = None,
        cell_len: int = 1,
    ) -> Generator[Tuple[list, int], None, None]:
        """使用 openpyxl 从 excel 的 sheet 中获取层次结构的表头

        Args:
            min_col (Union[int, str]): 范围，最小列，从1或A开始
            min_row (int): 范围，最小行，从1开始
            max_col (Union[int, str]): 范围，最大列，从1或A开始
            max_row (int): 范围，最大行，从1开始
            sheet (Worksheet): 工作表
            horizontal (bool, optional): 是否认为表头是从上到下的，否则是从左到右的
            root_list (List[list], optional): 递归时的根列表，无需传入
            cell_len (int, optional): 递归时的单元格长度/高度，无需传入

        Yields:
            Generator[Tuple[list, int], None, None]: 生成器，每次返回一个表头列表和单元格长度/高度
        """
        if isinstance(min_col, str):
            min_col = column_index_from_string(min_col)
        if isinstance(max_col, str):
            max_col = column_index_from_string(max_col)
        root_list = [] if root_list is None else root_list
        while min_col <= max_col and min_row <= max_row:
            cell: Cell = sheet.cell(row=min_row, column=min_col)
            merged_range = cls.find_merged_cell(sheet, cell)
            share_args = (sheet, horizontal, root_list + [cell.value])
            if merged_range:
                assert cell.column == merged_range.min_col, f"{cell.coordinate}, {merged_range.bounds}"
                assert min_row == merged_range.min_row, f"{cell.coordinate}, {merged_range.bounds}"
                if horizontal:
                    bounds = (min_col, merged_range.max_row + 1, merged_range.max_col, max_row)
                    yield from cls.get_table_header(*bounds, *share_args, merged_range.max_col - merged_range.min_col + 1)
                    min_col = merged_range.max_col + 1
                else:
                    bounds = (merged_range.max_col + 1, min_row, max_col, merged_range.max_row)
                    yield from cls.get_table_header(*bounds, *share_args, merged_range.max_row - merged_range.min_row + 1)
                    min_row = merged_range.max_row + 1
            else:
                if horizontal:
                    bounds = (min_col, min_row+1, min_col, max_row)
                    yield from cls.get_table_header(*bounds, *share_args, 1)
                    min_col += 1
                else:
                    bounds = (min_col+1, min_row, max_col, min_row)
                    yield from cls.get_table_header(*bounds, *share_args, 1)
                    min_row += 1
        if horizontal and min_row > max_row:
            yield root_list, cell_len
        elif not horizontal and min_col > max_col:
            yield root_list, cell_len

    @classmethod
    def get_docs_from_sheet(
        cls,
        sheet: Worksheet,
        th_end: Union[int, str],
        side_th_end: Union[int, str] = 0,
        horizontal: bool = True,
        start_col: Optional[Union[int, str]] = None,
        start_row: Optional[int] = None,
        end_col: Optional[Union[int, str]] = None,
        end_row: Optional[int] = None,
    ) -> dict:
        """从工作表中获取表格内容

        Args:
            sheet (Worksheet): 工作表
            th_end (Union[int, str]): 表头结束的位置，从1或A开始，horizontal为True时是数字行号，否则是字母列号
                比 start_col(horizontal=False) 或 start_row(horizontal=True) 小表示没有表头
            side_th_end (Union[int, str], optional): 侧边表头结束的位置，从1或A开始，horizontal为True时是字母列号，否则是数字行号
                比 start_col(horizontal=False) 或 start_row(horizontal=True) 小表示没有侧边表头
            horizontal (bool, optional): 表头是否是从上到下的，否则是从左到右的
            start_col (Optional[Union[int, str]], optional): 表格开始的列位置，从A开始，1也行，为None表示自动计算
            start_row (Optional[int], optional): 表格开始的行位置，从1开始，为None表示自动计算
            end_col (Optional[Union[int, str]], optional): 表格结束的列位置，从A开始，1也行，为None表示自动计算
            end_row (Optional[int], optional): 表格结束的行位置，从1开始，为None表示自动计算

        Returns:
            dict: 返回包括表格内容和侧边表头，表格内容是一个列表，每个元素是一条数据，侧边表头是一个嵌套字典
        """
        start_col = start_col or sheet.min_column
        start_row = start_row or sheet.min_row
        end_col = end_col or sheet.max_column
        end_row = end_row or sheet.max_row
        if isinstance(start_col, str):
            start_col = column_index_from_string(start_col)
        if isinstance(end_col, str):
            end_col = column_index_from_string(end_col)
        if isinstance(th_end, str):
            th_end = column_index_from_string(th_end)
        if isinstance(side_th_end, str):
            side_th_end = column_index_from_string(side_th_end)
        # 明确end下限
        if horizontal:
            th_end = max(th_end or 0, start_row - 1)
            has_th = th_end >= start_row
            side_th_end = max(side_th_end or 0, start_col - 1)
            has_side_th = side_th_end >= start_col
        else:
            th_end = max(th_end or 0, start_col - 1)
            has_th = th_end >= start_col
            side_th_end = max(side_th_end or 0, start_row - 1)
            has_side_th = side_th_end >= start_row
        # 获取表头
        if has_th:
            if horizontal:
                min_col, min_row, max_col, max_row = side_th_end + 1, start_row, end_col, th_end
            else:
                min_col, min_row, max_col, max_row = start_col, side_th_end + 1, th_end, end_row
            th_keys_len_L = list(cls.get_table_header(min_col, min_row, max_col, max_row, sheet, horizontal))
            th_len_D: NestedDict = pair_to_dict(th_keys_len_L)
            assert len(th_keys_len_L) == sum(1 for _ in dict_to_pair(th_len_D)), f"表头同一级可能存在同名现象: {th_len_D}"
        else:
            th_keys_len_L: List[Tuple[list, int]] = []
            th_len_D = {}
        # 获取侧边表头
        if has_side_th:
            if horizontal:
                min_col, min_row, max_col, max_row = start_col, th_end + 1, side_th_end, end_row
            else:
                min_col, min_row, max_col, max_row = th_end + 1, start_row, end_col, side_th_end
            side_th_keys_len_L = list(cls.get_table_header(min_col, min_row, max_col, max_row, sheet, not horizontal))
            side_th: NestedDict = pair_to_dict(side_th_keys_len_L)
            assert len(side_th_keys_len_L) == sum(1 for _ in dict_to_pair(side_th)), f"侧边表头同一级可能存在同名现象: {side_th}"
        else:
            side_th = {}
        # 获取表格内容
        td_L_L: List[list] = []  # 与 horizontal 无关的结果，1个list是一条数据
        if horizontal:
            for r in range(th_end + 1, end_row + 1):
                td_L = []
                for c in range(side_th_end + 1, end_col + 1):
                    td_L.append(sheet.cell(row=r, column=c).value)
                td_L_L.append(td_L)
        else:
            for c in range(th_end + 1, end_col + 1):
                td_L = []
                for r in range(side_th_end + 1, end_row + 1):
                    td_L.append(sheet.cell(row=r, column=c).value)
                td_L_L.append(td_L)
        # 生成 doc_L
        doc_L: List[dict] = []
        if has_th:
            for td_L in td_L_L:
                assert len(td_L) == len(th_keys_len_L), f"表格内容与表头长度不一致: {td_L}, {th_keys_len_L}"
                th_keys_v_L = []
                for v, (keys, l) in zip(td_L, th_keys_len_L):
                    assert l == 1, f"主表头最后一行不应该有合并单元格: {keys}, {l}"
                    th_keys_v_L.append((keys, v))
                doc_L.append(pair_to_dict(th_keys_v_L))
        return {
            'docs': doc_L,
            'side_th': side_th,
            'td_list': td_L_L,
            'th': th_len_D,
            'th_pairs': th_keys_len_L,
        }


if __name__ == "__main__":
    wb = load_workbook('test/openpyxl.xlsx')
    for name, th_end, side_th_end, horizontal in [
        ('横向', 3, 0, True),
        ('纵向', 'c', 0, False),
        ('横向2', 3, 3, True),
        ('纵向2', 'c', 3, False),
    ]:
        print('=' * 5, name, th_end, side_th_end, horizontal)
        sheet = wb[name]
        ret = XlsxReader.get_docs_from_sheet(sheet, th_end, side_th_end, horizontal=horizontal)
        pprint(ret)
        print()
