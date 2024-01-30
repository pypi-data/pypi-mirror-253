"""
@Project ：指纹识别 
@File    ：output_file.py
@IDE     ：PyCharm 
@Author  ：zhizhuo
@Date    ：2023/12/19 09:20 
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io


def get_excel(data):
    """
    将数据写到xlsx文件
    :param data:写入数据
    :return:二进制数据
    """
    illegal_chars = dict.fromkeys(range(0x00, 0x20))
    translator = str.maketrans(illegal_chars)
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['主机', 'URL', '协议', 'CMS', '标题', '状态码', '重定向次数',
               '服务器', '是否CDN', 'CDN IP列表', 'icon_hash', '证书']
    ws.append(headers)
    column_widths = [15, 20, 8, 20, 20, 8, 12, 20, 10, 20, 15, 15]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    green_fill = PatternFill(start_color="00FF00",
                             end_color="00FF00", fill_type="solid")
    for cell in ws[1]:
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for item in data:
        item['cdn_ip_list'] = ', '.join(item['cdn_ip_list'])
        if not item['cert']:
            item['cert'] = ''
        if item['is_cdn']:
            item['is_cdn'] = '是'
        else:
            item['is_cdn'] = '否'
        row = [str(value) for value in item.values()]
        row = ['' if str(value) == 'None' else value for value in row]
        row = [value.translate(translator) for value in row]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    return file_bytes.getvalue()
