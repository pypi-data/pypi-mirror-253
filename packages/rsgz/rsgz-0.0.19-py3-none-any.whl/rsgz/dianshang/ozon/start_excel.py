import openpyxl, os, time
from openpyxl import Workbook,load_workbook
from json_rich import json_rich_list
import settings
from settings import main_photo1, list_to_n_bei, size_list




def mk_xlsx(xlsx):
    import shutil
    if os.path.exists(xlsx):
        os.remove(xlsx)
    time.sleep(1)
    # openpyxl.Workbook().save(xlsx)
    shutil.copyfile("001.xlsx", xlsx)

def creat_excel():
    wb = load_workbook(settings.load_excel)  # 只读模式
    ws1 = wb["供应商模板"]
    ws2 = wb["PDF 文件"]
    ws3 = wb["Ozon.视频"]
    ws4 = wb["臭氧视频封面"]


    def write_d(hang1, hang2, lie, v):
        r"""
        适合写 一列 同一个值
        """
        n_sku=len(settings.size_list)*len(settings.color_list)
        for i in range(hang1, hang2+1):
            ws1.cell(row=i, column=lie).value=v

    n_sku = len(settings.size_list) * len(settings.color_list)
    for s in range(0+4,len(settings.sku_l)+4):
        ws1.cell(row=s, column=2).value = settings.sku_l[s-4]  # # 货号

    write_d(hang1=4, hang2=4 + n_sku-1, lie=3, v=settings.title)  # 商品名称
    write_d(hang1=4, hang2=4 + n_sku-1, lie=4, v=settings.price)  # 价格
    write_d(hang1=4, hang2=4 + n_sku-1, lie=5, v=settings.price_pre)  # 折前价格
    write_d(hang1=4, hang2=4 + n_sku-1, lie=6, v=settings.tax)  # 增值税
    write_d(hang1=4, hang2=4 + n_sku-1, lie=8, v=settings.type1)  # 商用型
    write_d(hang1=4, hang2=4 + n_sku-1, lie=10, v=settings.package_weight)  # 包装重量
    write_d(hang1=4, hang2=4 + n_sku-1, lie=11, v=settings.package_width)  # 包装宽度
    write_d(hang1=4, hang2=4 + n_sku-1, lie=12, v=settings.package_height)  # 包装高度
    write_d(hang1=4, hang2=4 + n_sku-1, lie=13, v=settings.package_length)  # 包装长度

    n_sku = len(settings.size_list) * len(settings.color_list)
    for s in range(0+4, n_sku+4):
        ws1.cell(row=s, column=14).value = settings.main_photo2[s-4]  # 链接主要照片
        ws1.cell(row=s, column=15).value = settings.other_photo[s-4]  # 链接主要照片

    write_d(hang1=4, hang2=4 + n_sku-1, lie=18, v=settings.brand)  # 服装和鞋类品牌
    write_d(hang1=4, hang2=4 + n_sku-1, lie=19, v=settings.card)  # 合并为一张卡片
    write_d(hang1=4, hang2=4 + n_sku-1, lie=20, v=settings.product_color1)  # 商品颜色

    size1_new = settings.size1_l*len(settings.color_list)
    size2_new = settings.size2_l*len(settings.color_list)
    e_wen_color_l = settings.product_color2_l_e_wen2

    for s in range(0+4,len(size1_new)+4):
        ws1.cell(row=s, column=21).value = size1_new[s-4]  # 俄罗斯尺码
        ws1.cell(row=s, column=22).value = size2_new[s-4]  # 由制造商规定尺码
        ws1.cell(row=s, column=23).value = e_wen_color_l[s-4]  # 俄文颜色
        ws1.cell(row=s, column=67).value = json_rich_list[s-4]  # json rich

    write_d(hang1=4, hang2=4 + n_sku-1, lie=24, v=settings.type2)  # 类型
    write_d(hang1=4, hang2=4 + n_sku-1, lie=25, v=settings.gender)  # 性别
    write_d(hang1=4, hang2=4 + n_sku-1, lie=26, v=settings.key_words)  # 关键字
    write_d(hang1=4, hang2=4 + n_sku-1, lie=27, v=settings.TargetAudience)  # 目标受众
    write_d(hang1=4, hang2=4 + n_sku-1, lie=28, v=settings.season)  # 季节
    write_d(hang1=4, hang2=4 + n_sku-1, lie=29, v=settings.model_height)  # 照片中的模特儿身高
    write_d(hang1=4, hang2=4 + n_sku-1, lie=30, v=settings.model_measurements)  # 照片中的模特儿的三围
    write_d(hang1=4, hang2=4 + n_sku-1, lie=31, v=settings.cloth_size)  # 图片上商品的尺码
    write_d(hang1=4, hang2=4 + n_sku-1, lie=32, v=settings.collect)  # 收集
    write_d(hang1=4, hang2=4 + n_sku-1, lie=33, v=settings.country_of_manufacture)  # 制造国
    write_d(hang1=4, hang2=4 + n_sku-1, lie=34, v=settings.print_type)  # 印花种类
    write_d(hang1=4, hang2=4 + n_sku-1, lie=35, v=settings.comments)  # 注解
    write_d(hang1=4, hang2=4 + n_sku-1, lie=36, v=settings.care_instructions)  # 保养说明
    # write_d(hang1=4, hang2=4 + n_sku-1, lie=37, v=settings.package_width)  # 服装和鞋类系列
    write_d(hang1=4, hang2=4 + n_sku-1, lie=38, v=settings.material)  # 材料
    write_d(hang1=4, hang2=4 + n_sku-1, lie=39, v=settings.material_ingredient)  # 材料成分
    write_d(hang1=4, hang2=4 + n_sku-1, lie=40, v=settings.gasket_inner_material)  # 垫片/内部材料
    write_d(hang1=4, hang2=4 + n_sku-1, lie=41, v=settings.filler)  # 填充材料
    write_d(hang1=4, hang2=4 + n_sku-1, lie=43, v=settings.temperature_range)  # 温度范围， °С
    write_d(hang1=4, hang2=4 + n_sku-1, lie=45, v=settings.style)  # 风格
    write_d(hang1=4, hang2=4 + n_sku-1, lie=46, v=settings.type_of_exercise)  # 运动种类
    write_d(hang1=4, hang2=4 + n_sku-1, lie=47, v=settings.clothing_type)  # 服装类型
    write_d(hang1=4, hang2=4 + n_sku-1, lie=48, v=settings.button_type)  # 扣子类型
    write_d(hang1=4, hang2=4 + n_sku-1, lie=53, v=settings.waist)  # 腰
    write_d(hang1=4, hang2=4 + n_sku-1, lie=58, v=settings.clothing_package_type)  # 服装包装类型
    write_d(hang1=4, hang2=4 + n_sku-1, lie=66, v=settings.JSON_size_table)  # JSON 大小表
    # write_d(hang1=4, hang2=4 + n_sku-1, lie=67, v=settings.JSON_rich_content)  # JSON 丰富内容
    write_d(hang1=4, hang2=4 + n_sku-1, lie=71, v=settings.breast_support_level)  # 乳房支撑水平

    # 处理第二个子表的 数据

    # 处理第三个 子表数据
    for s in range(0+4,len(settings.sku_l)+4):
        ws3.cell(row=s, column=1).value = settings.sku_l[s-4]  # 货号
        ws3.cell(row=s, column=2).value = settings.sku_l[s-4]+"_video"  # Ozon.视频: 名称
        ws3.cell(row=s, column=3).value = settings.video_url_list[s-4]  # Ozon视频: 链接
        ws3.cell(row=s, column=4).value = settings.sku_l[s-4]  # 视频中出现的产品 就是货号

        # 处理第四个 子表数据
        ws4.cell(row=s, column=1).value = settings.sku_l[s-4]  # 货号
        ws4.cell(row=s, column=2).value = settings.fengmian_url[s-4]  # 货号
    wb.save(settings.xlsx)


if __name__ == '__main__':
    mk_xlsx(xlsx=settings.xlsx)  # 新建
    creat_excel()
    print("表格 生成完毕！！！\n"*5)